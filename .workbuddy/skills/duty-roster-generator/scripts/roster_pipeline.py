# -*- coding: utf-8 -*-
"""主控流水线：
读取 人员 / 过往三天 -> 算未来3工作日(跳周一) -> 合并不可用(变更+补休+自动调休)
-> 求解 -> 成功则调 Excel COM 写入3份.xls并做格式指纹校验；
失败则只产《排班说明与告警》+《运行摘要》并给出3套方案。
所有重活都在本地脚本，对话层只需读取 运行摘要.json，token 消耗最低。
"""
import os
import sys
import json
import shutil
import datetime
import subprocess

_HERE = os.path.dirname(os.path.abspath(__file__))
_LIB = os.path.join(os.path.dirname(_HERE), ".lib")
if _LIB not in sys.path:
    sys.path.insert(0, _LIB)
import xlrd  # noqa: E402

from io_roster import (parse_personnel, parse_history, parse_buxiu,  # noqa: E402
                       find_project_root)
from scheduler import run_schedule, compute_future_dates  # noqa: E402
from template_text import render_cell, render_date, rebuild_note  # noqa: E402

WEEKDAY_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def locate_template(template_path, tmap):
    # formatting_info=True：空白模板 usedRange 可能塌缩为 A1:A19，
    # 不读取完整格式信息会漏掉 B-G 列，导致定位越界。
    book = xlrd.open_workbook(template_path, formatting_info=True)
    ws = book.sheet_by_name(tmap["sheet"])
    row_labels = {}
    for r in range(ws.nrows):
        v = ws.cell_value(r, 0)
        if isinstance(v, str) and v.strip():
            row_labels[r] = v.strip()
    located = {}
    for d in tmap["demands"]:
        found = next((r for r, lab in row_labels.items() if d["match"] in lab), None)
        if found is None:
            raise RuntimeError(f"模板找不到岗位行：{d['match']}")
        located[d["key"]] = f"{d['col']}{found + 1}"
    # 日期单元格（含日期模式的 A 列单元格）
    date_cell = tmap.get("date_cell")
    note_cell = tmap.get("note_cell")
    note_text = ""
    for r, lab in row_labels.items():
        import re
        if re.search(r"\d+月\d+日周[一二三四五六日]", lab):
            date_cell = f"A{r + 1}"
        if lab.startswith("其他"):
            note_cell = f"A{r + 1}"
            # 读取该备注行完整文本
            parts = []
            for c in range(ws.ncols):
                cv = ws.cell_value(r, c)
                if isinstance(cv, str) and cv.strip():
                    parts.append(cv)
            note_text = " ".join(parts)
    # 闸机关馆固定尾部（闭馆安保说明，常量，非排班池人员）—— 取自 template_map 固化值
    tail = tmap.get("gate_tail", "")
    # 排前询问岗（公务讲解/大屏制作/学宫）：定位单元格，由对话层决定是否排
    inquiry_cells = {}
    for ip in tmap.get("inquiry_posts", []):
        found = next((r for r, lab in row_labels.items() if ip["match"] in lab), None)
        if found is None:
            raise RuntimeError(f"模板找不到询问岗行：{ip['match']}")
        inquiry_cells[ip["match"]] = f"{ip['col']}{found + 1}"
    return located, date_cell, note_cell, note_text, tail, inquiry_cells


def day_kinds(changes, buxiu, rest_by_day, future):
    out = {}
    for fd in future:
        iso = fd["date"].isoformat()
        rest = rest_by_day.get(iso, [])
        leave, bx, sick, annual = [], [], [], []
        for u in changes.get("unavailable", []):
            if u.get("date") == iso:
                nm = u["name"]
                kind = u.get("kind", "请假")
                if kind in ("请假", "休假"):
                    leave.append(nm)
                elif kind == "补休":
                    bx.append(nm)
                elif kind == "病假":
                    sick.append(nm)
                elif kind == "年休假":
                    annual.append(nm)
        for nm, days in buxiu.items():
            if fd["date"].day in days and nm not in bx:
                bx.append(nm)
        out[iso] = {"rest": rest, "leave": leave, "buxiu": bx,
                    "sick": sick, "annual": annual}
    return out


def write_summary(batch, res, template_path, future, verified=None):
    summary = {
        "status": res["status"],
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "template": os.path.basename(template_path),
        "future_dates": [f"{fd['date'].isoformat()}({fd['weekday']})" for fd in future],
        "days": [
            {
                "date": d["date"], "weekday": d["weekday"],
                "rest": d["rest"],
                "unavailable": d["unavailable"],
                "assignments": d["assignments"],
                "issues": d["issues"],
            } for d in res["days"]
        ],
        "issues": res["issues"],
        "options": res["options"],
    }
    if verified is not None:
        summary["format_verification"] = "passed" if verified else "FAILED"
    with open(os.path.join(batch, "运行摘要.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    return summary


def write_report(batch, res, template_path, future, located, tail, kinds,
                 personnel, history, changes, verified=None, inquiry_report=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    bold = Font(bold=True)
    ws.append(["镇海楼值班表 · 自动排班说明与告警"])
    ws.append([])
    ws.append(["生成时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["复制模板", os.path.basename(template_path)])
    ws.append(["排班状态", "✅ 已生成可发布" if res["status"] == "ok" else "⛔ 已阻断（未发布日表）"])
    ws.append(["未来三天", "、".join(f"{fd['date'].isoformat()}({fd['weekday']})" for fd in future)])
    ws.append([])
    # 排前询问岗决策
    if inquiry_report:
        ws.append(["排前询问岗（本轮是否安排）"])
        for m, want in inquiry_report.items():
            ws.append([f"  {m}", "已排（待负责人指派专人）" if want else "未安排（按惯例留空）"])
        ws.append([])
    for d in res["days"]:
        k = kinds.get(d["date"], {})
        ws.append([f"【{d['date']} {d['weekday']}】"])
        ws.append(["  调休", "、".join(d["rest"]) or "—"])
        ws.append(["  请假/补休/病假", "、".join(k.get("leave", []) + k.get("buxiu", []) +
                                                k.get("sick", []) + k.get("annual", [])) or "—"])
        if d["issues"]:
            ws.append(["  告警", "；".join(d["issues"])])
    if verified is not None:
        ws.append([])
        ws.append(["格式指纹校验", "通过" if verified else "未通过（见 diff）"])
    # 方案页
    if res["options"]:
        ws2 = wb.create_sheet("告警与方案")
        ws2.append(["不可行原因"])
        for i in res["issues"]:
            ws2.append(["- " + i])
        ws2.append([])
        ws2.append(["建议方案（优缺点 / 沟通对象 / 话术）"])
        for opt in res["options"]:
            ws2.append([opt["title"]])
            ws2.append(["优点", opt["pros"]])
            ws2.append(["缺点", opt["cons"]])
            ws2.append(["沟通对象", opt["communicate_with"]])
            ws2.append(["建议话术", opt["suggested_wording"]])
            ws2.append([])
    else:
        ws2 = wb.create_sheet("排班预览")
        ws2.append(["日期", "岗位", "人员"])
        for d in res["days"]:
            for key, names in d["assignments"].items():
                ws2.append([d["date"], key, "、".join(names) if names else "（空缺）"])
    wb.save(os.path.join(batch, "排班说明与告警.xlsx"))


def run(changes, base):
    """生成未来 3 天排班。

    changes 结构：
      { "unavailable": [{name,date,kind,scope}],   # 请假/休假/病假/年休假/补休
        "rest":        [{name,date}],              # 用户强制调休
        "inquiry":     {"公务讲解": true, ...} }    # 排前询问岗是否本轮安排
    若 inquiry 缺省，则所有询问岗默认不安排（安全），并在报告中提示待确认。
    """
    from rest_state import load_or_seed, plan_rest, save as save_state  # noqa

    personnel = parse_personnel(os.path.join(base, "输入", "人员", "人员.xls"))
    history = parse_history(os.path.join(base, "输入", "过往三天"), personnel)
    if not history:
        raise RuntimeError("过往三天文件夹没有可解析的值班表")

    # 空白标准模板（每次复制它、只填白名单格，保证所有表格位置永久对齐）
    tpl_path = os.path.join(os.path.dirname(_HERE), "templates", "标准空白值班表.xls")
    if not os.path.exists(tpl_path):
        raise RuntimeError(f"空白标准模板缺失：{tpl_path}")
    # 模板日期用于推算未来 3 天：取最新历史文件的日期
    template_date = datetime.date.fromisoformat(history[-1]["date"])
    buxiu = parse_buxiu(os.path.join(base, "输入", "过往三天",
                                     os.path.basename(history[-1]["path"])), personnel)

    with open(os.path.join(_HERE, "template_map.json"), encoding="utf-8") as f:
        tmap = json.load(f)
    located, date_cell, note_cell, note_text, tail, inquiry_cells = locate_template(tpl_path, tmap)
    future = compute_future_dates(template_date, 3)

    # —— 调休轮转：读状态 -> 排 ~3 人/日 -> 成功后回写 ——
    state = load_or_seed(None, history)
    names = [p["name"] for p in personnel]
    tags_of = {p["name"]: set(p["tags"]) for p in personnel}
    unavail_by_day = {}
    for fd in future:
        iso = fd["date"].isoformat()
        s = set()
        for u in changes.get("unavailable", []):
            if u.get("date") == iso:
                s.add(u["name"])
        for nm, days in buxiu.items():
            if fd["date"].day in days:
                s.add(nm)
        unavail_by_day[iso] = s
    # 调休频率：按历史日均浮动（1-5），非固定 3
    rest_counts = [len(h.get("rest", [])) for h in history]
    daily_target = int(round(sum(rest_counts) / len(rest_counts))) if rest_counts else 3
    daily_target = max(1, min(5, daily_target))
    rest_by_day = plan_rest(state, names, future, unavail_by_day,
                            forced={u["date"]: {u["name"]} for u in changes.get("rest", [])},
                            tags_of=tags_of, daily_target=daily_target)

    res = run_schedule(personnel, history, tmap["demands"], changes, template_date,
                       future, buxiu=buxiu, rest_by_day=rest_by_day)
    kinds = day_kinds(changes, buxiu, res["rest_by_day"], future)

    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    batch = os.path.join(base, "输出", stamp)
    os.makedirs(batch, exist_ok=True)

    inquiry_dec = changes.get("inquiry") or {}
    inquiry_report = {}

    if res["status"] == "blocked":
        write_summary(batch, res, tpl_path, future)
        write_report(batch, res, tpl_path, future, located, tail, kinds,
                     personnel, history, changes)
        return {"status": "blocked", "batch": batch,
                "message": "无法满足硬约束，已阻断并给出3套方案。"}

    # 渲染 + 构造写入载荷
    days_payload = []
    for d in res["days"]:
        cells = []
        for dm in tmap["demands"]:
            key = dm["key"]
            names_assigned = d["assignments"].get(key, [])
            t = render_cell(key, names_assigned, tail if key == "闸机关馆" else "")
            cells.append({"cell": located[key], "text": t})
        # 排前询问岗：本轮确认要排的，写入占位标记待负责人指派
        for match, cell in inquiry_cells.items():
            want = inquiry_dec.get(match, False)
            inquiry_report[match] = bool(want)
            if want:
                cells.append({"cell": cell, "text": "（待安排·请负责人指派）"})
        k = kinds.get(d["date"], {})
        note = rebuild_note(note_text, k.get("rest", []), k.get("leave", []),
                            k.get("buxiu", []), k.get("sick", []), k.get("annual", []))
        days_payload.append({
            "filename": f"{d['date']}值班表.xls",
            "date_cell": date_cell, "date_text": render_date(d["date"], d["weekday"]),
            "note_cell": note_cell, "note_text": note, "cells": cells,
        })
    payload = {"template_path": os.path.abspath(tpl_path),
               "sheet": tmap["sheet"], "output_dir": os.path.abspath(batch),
               "days": days_payload}
    payload_path = os.path.join(batch, ".payload.json")
    with open(payload_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)

    ps = os.path.join(_HERE, "write_template.ps1")
    r = subprocess.run(["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass",
                        "-File", ps, "-PayloadPath", payload_path],
                       capture_output=True, text=True, encoding="utf-8",
                       errors="replace", timeout=500)
    if r.returncode != 0:
        write_summary(batch, res, tpl_path, future)
        with open(os.path.join(batch, "运行摘要.json"), "r", encoding="utf-8") as f:
            s = json.load(f)
        s["status"] = "error"
        s["error"] = (r.stderr or r.stdout).strip()
        with open(os.path.join(batch, "运行摘要.json"), "w", encoding="utf-8") as f:
            json.dump(s, f, ensure_ascii=False, indent=2)
        return {"status": "error", "batch": batch, "error": s["error"]}

    # 格式指纹校验（白名单 = 日期 + 备注 + 各需求格 + 本轮写入的询问岗格）
    allowed = [date_cell, note_cell] + list(located.values())
    for match, cell in inquiry_cells.items():
        if inquiry_report.get(match):
            allowed.append(cell)
    out_paths = ",".join(os.path.join(batch, dp["filename"]) for dp in days_payload)
    rep = os.path.join(batch, ".fingerprint.json")
    fps = os.path.join(_HERE, "fingerprint.ps1")
    rf = subprocess.run(["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass",
                         "-File", fps, "-TemplatePath", os.path.abspath(tpl_path),
                         "-OutputPaths", out_paths, "-AllowedCells", ",".join(allowed),
                         "-ReportPath", rep],
                        capture_output=True, text=True, encoding="utf-8",
                        errors="replace", timeout=500)
    verified = False
    if rf.returncode == 0 and os.path.exists(rep):
        repj = json.load(open(rep, encoding="utf-8-sig"))
        verified = repj.get("verified", False)
        if not verified:
            res["issues"].append("格式指纹校验未通过：" + "；".join(repj.get("diffs", [])[:8]))

    # 仅在「成功且指纹通过」时回写调休状态（阻断/失败不推进轮转）
    if verified:
        save_state(state)

    write_summary(batch, res, tpl_path, future, verified=verified)
    write_report(batch, res, tpl_path, future, located, tail, kinds,
                 personnel, history, changes, verified=verified,
                 inquiry_report=inquiry_report)
    daily = [dp["filename"] for dp in days_payload]
    return {"status": "ok", "batch": batch, "verified": verified, "daily": daily,
            "summary_path": os.path.join(batch, "运行摘要.json")}


if __name__ == "__main__":
    base = find_project_root(_HERE)
    changes = {"unavailable": [], "rest": []}
    # 可选：从命令行指定的变更 JSON 读取
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        with open(sys.argv[1], encoding="utf-8") as f:
            changes = json.load(f)
    result = run(changes, base)
    print(json.dumps(result, ensure_ascii=False, indent=2))
